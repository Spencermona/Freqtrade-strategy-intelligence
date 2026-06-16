from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / "results"
INPUT_CSV = RESULTS_DIR / "comprehensive_strategy_ranking.csv"

OUTPUT_FILES = {
    "zh": RESULTS_DIR / "Freqtrade策略评分汇总表.xlsx",
    "en": RESULTS_DIR / "freqtrade_strategy_report.xlsx",
}

TEXT = {
    "zh": {
        "top_sheet": "Top100策略",
        "summary_sheet": "统计摘要",
        "category_sheet": "维度评分",
        "timeframe_sheet": "周期对比",
        "risk_sheet": "风险等级分析",
        "strategy_name": "策略名称",
        "timeframe": "时间周期",
        "overall_score": "总体评分",
        "technical_indicators_score": "指标质量",
        "entry_logic_score": "入场逻辑",
        "exit_strategy_score": "出场策略",
        "risk_management_score": "风险管理",
        "code_quality_score": "代码质量",
        "market_adaptability_score": "市场适应性",
        "risk_assessment": "风险等级",
        "complexity_level": "复杂度",
        "estimated_sharpe_ratio": "夏普比率",
        "max_drawdown_estimate": "最大回撤",
        "metric": "指标",
        "value": "数值",
        "total": "总策略数",
        "average": "平均评分",
        "highest": "最高评分",
        "lowest": "最低评分",
        "pass": "及格率 >= 60分",
        "excellent": "优秀率 >= 70分",
        "elite": "卓越率 >= 80分",
        "risk_low": "低风险策略",
        "risk_medium": "中等风险策略",
        "risk_high": "高风险策略",
        "complexity_beginner": "初级复杂度",
        "complexity_intermediate": "中级复杂度",
        "complexity_advanced": "高级复杂度",
        "dimension": "维度",
        "avg": "平均分",
        "max": "最高分",
        "min": "最低分",
        "above8": ">8分策略",
        "below5": "<5分策略",
        "count": "策略数",
        "share": "占比",
        "avg_risk_return": "平均风险收益",
        "avg_drawdown": "平均回撤",
        "saved": "Excel 汇总表已生成",
        "final": "Freqtrade 策略分析 - 最终汇总",
        "median": "中位数",
        "std": "标准差",
        "top10": "Top 10 策略",
    },
    "en": {
        "top_sheet": "Top100 Strategies",
        "summary_sheet": "Summary",
        "category_sheet": "Category Scores",
        "timeframe_sheet": "Timeframe Comparison",
        "risk_sheet": "Risk Analysis",
        "strategy_name": "Strategy Name",
        "timeframe": "Timeframe",
        "overall_score": "Overall Score",
        "technical_indicators_score": "Technical Indicators",
        "entry_logic_score": "Entry Logic",
        "exit_strategy_score": "Exit Strategy",
        "risk_management_score": "Risk Management",
        "code_quality_score": "Code Quality",
        "market_adaptability_score": "Market Adaptability",
        "risk_assessment": "Risk Level",
        "complexity_level": "Complexity",
        "estimated_sharpe_ratio": "Estimated Sharpe",
        "max_drawdown_estimate": "Max Drawdown",
        "metric": "Metric",
        "value": "Value",
        "total": "Total Strategies",
        "average": "Average Score",
        "highest": "Highest Score",
        "lowest": "Lowest Score",
        "pass": "Pass Rate >= 60",
        "excellent": "Excellent Rate >= 70",
        "elite": "Elite Rate >= 80",
        "risk_low": "Low-Risk Strategies",
        "risk_medium": "Medium-Risk Strategies",
        "risk_high": "High-Risk Strategies",
        "complexity_beginner": "Beginner Complexity",
        "complexity_intermediate": "Intermediate Complexity",
        "complexity_advanced": "Advanced Complexity",
        "dimension": "Dimension",
        "avg": "Average",
        "max": "Max",
        "min": "Min",
        "above8": ">8 Score Strategies",
        "below5": "<5 Score Strategies",
        "count": "Strategy Count",
        "share": "Share",
        "avg_risk_return": "Avg Risk-Adjusted Return",
        "avg_drawdown": "Avg Drawdown",
        "saved": "Excel report generated",
        "final": "Freqtrade Strategy Analysis - Final Summary",
        "median": "Median",
        "std": "Standard Deviation",
        "top10": "Top 10 Strategies",
    },
}


def extract_drawdown_floor(value: object) -> int:
    if pd.isna(value) or value == "unknown":
        return 0
    try:
        return int(str(value).split("-")[0].replace("%", ""))
    except ValueError:
        return 0


def write_summary_workbook(df: pd.DataFrame, lang: str) -> Path:
    labels = TEXT[lang]
    output_xlsx = OUTPUT_FILES[lang]
    RESULTS_DIR.mkdir(exist_ok=True)
    top100 = df.nlargest(100, "overall_score").copy()

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        top_columns = [
            "strategy_name",
            "timeframe",
            "overall_score",
            "technical_indicators_score",
            "entry_logic_score",
            "exit_strategy_score",
            "risk_management_score",
            "code_quality_score",
            "market_adaptability_score",
            "risk_assessment",
            "complexity_level",
            "estimated_sharpe_ratio",
            "max_drawdown_estimate",
        ]
        top100_display = top100[top_columns].reset_index(drop=True)
        top100_display.columns = [labels[column] for column in top_columns]
        top100_display.to_excel(writer, sheet_name=labels["top_sheet"], index=False)

        summary_data = {
            labels["metric"]: [
                labels["total"],
                labels["average"],
                labels["highest"],
                labels["lowest"],
                labels["pass"],
                labels["excellent"],
                labels["elite"],
                "",
                "5m",
                "3m",
                "1m",
                "",
                labels["risk_low"],
                labels["risk_medium"],
                labels["risk_high"],
                "",
                labels["complexity_beginner"],
                labels["complexity_intermediate"],
                labels["complexity_advanced"],
            ],
            labels["value"]: [
                len(df),
                f"{df['overall_score'].mean():.2f}/100",
                f"{df['overall_score'].max()}/100",
                f"{df['overall_score'].min()}/100",
                f"{(df['overall_score'] >= 60).sum()} ({(df['overall_score'] >= 60).sum() / len(df) * 100:.1f}%)",
                f"{(df['overall_score'] >= 70).sum()} ({(df['overall_score'] >= 70).sum() / len(df) * 100:.1f}%)",
                f"{(df['overall_score'] >= 80).sum()} ({(df['overall_score'] >= 80).sum() / len(df) * 100:.1f}%)",
                "",
                len(df[df["timeframe"] == "5m"]),
                len(df[df["timeframe"] == "3m"]),
                len(df[df["timeframe"] == "1m"]),
                "",
                len(df[df["risk_assessment"] == "low"]),
                len(df[df["risk_assessment"] == "medium"]),
                len(df[df["risk_assessment"] == "high"]),
                "",
                len(df[df["complexity_level"] == "beginner"]),
                len(df[df["complexity_level"] == "intermediate"]),
                len(df[df["complexity_level"] == "advanced"]),
            ],
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name=labels["summary_sheet"], index=False)

        category_cols = [col for col in df.columns if col.endswith("_score") and col != "overall_score"]
        category_stats = []
        for col in category_cols:
            category_stats.append(
                {
                    labels["dimension"]: col.replace("_score", "").replace("_", " ").title(),
                    labels["avg"]: f"{df[col].mean():.2f}",
                    labels["max"]: f"{df[col].max()}",
                    labels["min"]: f"{df[col].min()}",
                    labels["above8"]: (df[col] > 8).sum(),
                    labels["below5"]: (df[col] < 5).sum(),
                }
            )
        pd.DataFrame(category_stats).to_excel(writer, sheet_name=labels["category_sheet"], index=False)

        timeframe_stats = []
        for tf in df["timeframe"].dropna().unique():
            tf_data = df[df["timeframe"] == tf]
            timeframe_stats.append(
                {
                    labels["timeframe"]: tf,
                    labels["count"]: len(tf_data),
                    labels["avg"]: f"{tf_data['overall_score'].mean():.2f}",
                    labels["max"]: f"{tf_data['overall_score'].max()}",
                    labels["min"]: f"{tf_data['overall_score'].min()}",
                    labels["pass"]: f"{(tf_data['overall_score'] >= 60).sum() / len(tf_data) * 100:.1f}%",
                    labels["excellent"]: f"{(tf_data['overall_score'] >= 70).sum() / len(tf_data) * 100:.1f}%",
                }
            )
        pd.DataFrame(timeframe_stats).to_excel(writer, sheet_name=labels["timeframe_sheet"], index=False)

        risk_analysis = []
        for risk in df["risk_assessment"].dropna().unique():
            risk_data = df[df["risk_assessment"] == risk]
            risk_analysis.append(
                {
                    labels["risk_assessment"]: risk,
                    labels["count"]: len(risk_data),
                    labels["share"]: f"{len(risk_data) / len(df) * 100:.1f}%",
                    labels["avg"]: f"{risk_data['overall_score'].mean():.2f}",
                    labels["avg_risk_return"]: f"{risk_data['risk_adjusted_returns_score'].mean():.2f}",
                    labels["avg_drawdown"]: f"{risk_data['max_drawdown_estimate'].apply(extract_drawdown_floor).mean():.0f}%",
                }
            )
        pd.DataFrame(risk_analysis).to_excel(writer, sheet_name=labels["risk_sheet"], index=False)

    return output_xlsx


def format_workbook(output_xlsx: Path) -> None:
    wb = openpyxl.load_workbook(output_xlsx)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(output_xlsx)


def print_console_summary(df: pd.DataFrame, lang: str, output_xlsx: Path) -> None:
    labels = TEXT[lang]
    print(f"{labels['saved']}: {output_xlsx}")
    print("\n" + "=" * 70)
    print(labels["final"])
    print("=" * 70)
    print(f"{labels['total']}: {len(df)}")
    print(f"{labels['average']}: {df['overall_score'].mean():.2f}/100")
    print(f"{labels['median']}: {df['overall_score'].median():.2f}/100")
    print(f"{labels['std']}: {df['overall_score'].std():.2f}")
    print(f"\n{labels['top10']}:")
    columns = ["strategy_name", "overall_score", "timeframe", "risk_assessment"]
    for i, (_, row) in enumerate(df.nlargest(10, "overall_score")[columns].iterrows(), 1):
        print(f"{i:2d}. {row['strategy_name']:<30s} {row['overall_score']:5.0f} ({row['timeframe']}, {row['risk_assessment']})")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate a localized Freqtrade strategy Excel report.")
    parser.add_argument("--language", choices=["zh", "en"], default="en")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if not INPUT_CSV.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_CSV}")
    df = pd.read_csv(INPUT_CSV)
    output_xlsx = write_summary_workbook(df, args.language)
    format_workbook(output_xlsx)
    print_console_summary(df, args.language, output_xlsx)


if __name__ == "__main__":
    main()
